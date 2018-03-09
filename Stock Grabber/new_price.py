import openpyxl, yahoo_finance, urllib.request, json


wb = openpyxl.load_workbook('stock_values.xlsx')
sheet = wb.get_sheet_by_name('AwesomeSheet')
stocks = ['AAPL', 'BABA', 'GPRO', 'EA', 'ATVI', 'VFINX', 'ETH']
stocks_length = len(stocks)

#Stuff for ether data
with urllib.request.urlopen("https://api.coinmarketcap.com/v1/ticker/bitcoin/") as url:
    ether_data = json.loads(url.read().decode())
    daily_eth_price = ether_data["price"]["usd"]
    daily_eth_percent_change = ether_data["change"]

for i in range(stocks_length):
    if stocks[i] != 'ETH':
        stock = yahoo_finance.Share(stocks[i])
        sheet.cell(row=3 + i, column=7).value = float(stock.get_prev_close())
        sheet.cell(row=3 + i, column=9).value = stock.get_percent_change()
        print('Your ' + stocks[i] + ' stock has changed by: ' + str(stock.get_percent_change()))
    else:
        sheet.cell(row=3 + i, column=7).value = float(daily_eth_price)
        #Had to add a silly fix to ETH since excel pulls it in as 211% instead of 2.11%
        sheet.cell(row=3 + i, column=9).value = float(daily_eth_percent_change) * 0.01
        print('Your ' + stocks[i] + ' stock has changed by: ' + str(stock.get_percent_change()))

print("Your stocks were updated!")
wb.save("stock_values.xlsx")
