import openpyxl

wb = openpyxl.load_workbook('Copy of Sales ops.xlsx')
sheet = wb.get_sheet_by_name('Sheet2')
output_sheet = wb.create_sheet('indi_commas')

#Add commas to the values from sheet, then store them in the output_sheet
for i in range(1, 4996):
    output_sheet.cell(row=i, column=1).value = str(sheet.cell(row=i, column=2).value) + ', '

wb.save('Copy of Sales ops.xlsx')




individual_id = 99999

if individual_id > 100000:
    print("Individual is a Group Leader")
else:
    print("Individual is a Student")
